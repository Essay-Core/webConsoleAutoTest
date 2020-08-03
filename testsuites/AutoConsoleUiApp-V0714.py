# 标准库
import sys
import os
import time
import configparser
import threading
import _thread
from selenium import webdriver
from datetime import datetime
from selenium.webdriver.chrome.options import Options

# 第三方库
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox,QFileDialog
from PyQt5.QtWidgets import QTableWidgetItem, QAbstractItemView, QHeaderView
from PyQt5.QtGui import QFont, QBrush, QColor
from PyQt5.QtCore import Qt,QDateTime
from PyQt5.QtCore import QThread, pyqtSignal

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font

# 项目库
from ui.table import Ui_MainWindow
from framework.logger import Logger

# 单元格颜色填充选项
red_fill = PatternFill("solid", fgColor="FF0000")
green_fill = PatternFill("solid", fgColor="00FF00")

# 字体格式，颜色和大小
font_pass = Font(size=16, bold=True, color="00FF00")
font_false = Font(size=16, bold=True, color="FF0000")

online_color = QBrush(QColor(95, 215, 169))
offline_color = QBrush(QColor(255, 0, 0))

ip_online_dictionary = {}  # 字典
g_ip_config = [""] * 120  # 记录读取的配置文件ip   60 -》120

lock = threading.Lock()  # 创建Lock对象
pid_ip_dictionary = {}  # 进程号：IP
pid_div_id_dictionary = {}  # 进程号:  设备ID
pa_opera_dict = {}

# 获取系统当前时间,不会改变
stime = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
logger = Logger(logger='MyTestV2').getlog()
filename = stime

Chrome_path = r"../bin/chromedriver.exe"

# 先定位主筐体的位置，让后通过相对位置来定位表格中的元素
main_xpath = '//main[@class="el-main main"]'
url = "http://192.168.3.100/#/login"

inputuser_box = "//input[@placeholder='请输入用户名']"
inputpasswd_box = "//input[@type='password']"
login_submit_btn = "//*[@id='app']/div/div/form/button"  # 登录按钮

# 主控重启
reboot_mcband = '//button[@class="el-button el-button--danger el-button--mini is-round"]'  # 重启系统
comfirm = '//button[@class="el-button el-button--default el-button--small el-button--primary "]'

band_1 = "//div[text()='1']"  # 列表中第1块基带板
band_2 = "//div[text()='2']"  # 列表中第2块基带板
band_3 = "//div[text()='3']"  # 列表中第3块基带板

setpar_band = "//span[text()='参数设置']"
UePMax = "//*[@id='app']/section/section/main/div/div[2]/div[3]/div[4]/div/div[2]/form/div[13]/div[2]/div/div/div/div/input"  # 输出功率
save_par = "//span[text()='保存参数']"
comfirm_button = "//button[@class='el-button el-button--default el-button--small el-button--primary ']"

jz_manage = "//span[text()='基站管理']"
realtimedata = "//span[text()='实时数据']"
kb_temperature = "//span[text()='核心板温度']"
band_manage = "//span[text()='核心板管理']"
pa_manage = "//span[text()='功放管理']"

reboot_band = "//span[text()='软重启']"
power_outage_restart = "//span[text()='断电重启']"

pa_on = "//span[text()='开启所有功放']"
pa_off = "//span[text()='关闭所有功放']"
refresh = "//i[@class='el-icon-refresh']"
warning = '//div[@class="el-message-box__status el-icon-warning"]'

# B40
B40_table_pa_xpath = "//*[@id='app']/section/section/main/div/div[2]/div[3]/div[2]/div[3]/table/tbody/tr[1]/td[9]"
B40_table_pa_css = "#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(1) > td.el-table_1_column_9 > div > div > span"

# 读取B1功放
B1_table_pa_xpath = "//*[@id='app']/section/section/main/div/div[2]/div[3]/div[2]/div[3]/table/tbody/tr[2]/td[9]"
B1_table_pa_css = "#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(2) > td.el-table_1_column_9 > div > div > span"

# 读取gsm功放
GSM_table_pa_xpath = "//*[@id='app']/section/section/main/div/div[2]/div[3]/div[2]/div[3]/table/tbody/tr[3]/td[9]"
GSM_table_pa_css = "#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(3) > td.el-table_1_column_9 > div > div > span"

# 核心板: OS：online status ;WS: work status;
# FDD-56
OS_FDD_56_xpath_iframe = '//*[@id="app"]/section/section/main/div/div[2]/div[4]/div[2]/div[3]/table/tbody/tr[1]/td[7]/div/span'
OS_FDD_56_css_iframe = ''
OS_FDD_56_xpath = '//*[@id="app"]/section/section/main/div/div[2]/div[4]/div[2]/div[3]/table/tbody/tr[1]/td[7]'
OS_FDD_56_css = '#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(1) > td.el-table_6_column_105 > div > span'
WS_FDD_56_xpath = '//*[@id="app"]/section/section/main/div/div[2]/div[4]/div[2]/div[3]/table/tbody/tr[1]/td[8]'
WS_FDD_56_css = '#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(1) > td.el-table_6_column_106 > div > div > span'

# TDD-55
OS_TDD_55_xpath = '//*[@id="app"]/section/section/main/div/div[2]/div[4]/div[2]/div[3]/table/tbody/tr[2]/td[7]'
OS_TDD_55_css = '#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr.el-table__row.expanded > td.el-table_6_column_105 > div > span'
WS_TDD_55_xpath = '//*[@id="app"]/section/section/main/div/div[2]/div[4]/div[2]/div[3]/table/tbody/tr[2]/td[8]'
WS_TDD_55_css = '#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr.el-table__row.expanded > td.el-table_6_column_106 > div > div > span'

# OnlineStatus
OS_GSM_KB_xpath = '//*[@id="app"]/section/section/main/div/div[2]/div[4]/div[2]/div[3]/table/tbody/tr[4]/td[7]'
OS_GSM_KB_css = '#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(4) > td.el-table_6_column_105 > div > span'
WS_GSM_KB_xpath = '//*[@id="app"]/section/section/main/div/div[2]/div[4]/div[2]/div[3]/table/tbody/tr[4]/td[8]'
WS_GSM_KB_css = '#app > section > section > main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(4) > td.el-table_6_column_106 > div > div > span'

# 测试结果
resultReportHeadLine = [
    "设备IP",
    "设备ID",
    "B40功放关闭输出功率",  # 输出功率
    "B1功放关闭输出功率",
    "GSM功放关闭输出功率",
    "B40功放开启输出功率",
    "B1功放开启输出功率",
    "GSM功放开启输出功率",

    "功放关闭时间",  # 关闭
    "B40功放关闭",
    "B1功放关闭",
    "GSM功放关闭",

    "功放开启时间",
    "B40功放开启",  # 开启
    "B1功放开启",
    "GSM功放开启",

    "核心板1软重启时间",  # 软重启1
    "核心板1软重启建立小区时间",
    "核心板1软重启",

    "核心板2软重启时间",  # 软重启2
    "核心板2软重启建立小区时间",
    "核心板2软重启",

    "核心板3软重启时间",  # 软重启3
    "核心板3软重启建立小区时间",
    "核心板3软重启",

    "核心板1断电重启时间",  # 断电重启1
    "核心板1断电重启建立小区时间",
    "核心板1断电重启",

    "核心板2断电重启时间",  # 断电重启2
    "核心板2断电重启建立小区时间",
    "核心板2断电重启",

    "核心板3断电重启时间",  # 断电重启3
    "核心板3断电重启建立小区时间",
    "核心板3断电重启",

    "核心板类型",
    "核心板温度",
    "核心板类型",
    "核心板温度",
    "核心板类型",
    "核心板温度",

    "主控板重启时间",  # 主控重启
    "主控板上线时间",
    "主控板重启",
    "测试结果"
]

# 测试结果，超讯
resultReportHeadLine_CX = [
    "设备IP",
    "设备ID",

    "核心板1软重启时间",  # 软重启1
    "核心板1软重启建立小区时间",
    "核心板1软重启",

    "核心板2软重启时间",  # 软重启2
    "核心板2软重启建立小区时间",
    "核心板2软重启",

    "核心板3软重启时间",  # 软重启3
    "核心板3软重启建立小区时间",
    "核心板3软重启",

    "主控板重启时间",  # 主控重启
    "主控板上线时间",
    "主控板重启",
    "测试结果"
]

# 测试结果，百才帮
resultReportHeadLine_BCB = [
    "设备IP",
    "设备ID",

    "核心板1断电重启时间",  # 断电重启1
    "核心板1断电重启建立小区时间",
    "核心板1断电重启",

    "核心板2断电重启时间",  # 断电重启2
    "核心板2断电重启建立小区时间",
    "核心板2断电重启",

    "核心板3断电重启时间",  # 断电重启3
    "核心板3断电重启建立小区时间",
    "核心板3断电重启",

    "主控板重启时间",  # 主控重启
    "主控板上线时间",
    "主控板重启",
    "测试结果"
]

# 测试结果
MCB_REBOOT_HEADLINE = [
    "设备IP",
    "设备ID",
    "主控板重启时间",  # 主控重启
    "主控板上线时间",
    "主控板重启",
    "测试结果"
]

# 测试结果
GPIO_POWER_DOWN_REBOOT_HEADLINE = [
    "设备IP",
    "设备ID",
    "核心板重启时间",
    "核心板重启上线时间",
    "核心板重启",
    "测试结果"
]

LOGIN_USER = 'admin'
LOGIN_PASSWD = '123456'

################应用接口上##########################
'''
函数名：PA_Reboot
功能：
    打开功放，记录时间，查询功放输出功率
    关闭功放，记录时间，查询功放输出功率
流程：
    1，初始化参数
    1.2 登录
    2，设置输出功率
    3，关闭所有功放
    4，查询输出功率
    5，开启功放
    6，查询输出功率
    7，返回
err:
    1,打开url失败
    2，定位元素失败
参数：
    url:地址，如http://192.168.3.200/#/login
返回值：
    paRetStatus = [ '时间','false','false','false',
                    '时间','false','false','false']
    err_flags:
        1:成功
        -1:失败
修改时间：
    2020年4月29日09:59:53
    2020年4月30日14:04:37
        不通过功放输出功率判断开关结果，只需记录功放值 --ok--ok
        通过按钮状态判断是否开关--
        增加刷新后的延迟时间，5s --ok
        返回值添加功放值列 --ok
    2020年5月5日11:47:32
        添加返回值，错误标志
    2020年5月8日10:21:31
        添加属性开关值得断言定位获取
'''


def pa_warning_detect(driver):
    try:
        # printf(warning)
        driver.find_element_by_xpath(warning)
        # printf(comfirm)
        driver.find_element_by_xpath(comfirm).click()
        time.sleep(1)
        printf("查找元素成功")
        return True
    except Exception as err:
        printf(err)
        return False


def PA_Reboot(url):
    paRetStatus = ['', 'false', 'false', 'false',
                   '', 'false', 'false', 'false']
    paCheckValus = ['-999'] * 6
    indexNu = 0
    driver, loginStatus = login(url)
    if loginStatus == 1:
        printf("登录成功啦")
    else:
        printf("登录失败，等待10s后再次尝试登录")
        time.sleep(10)
        driver, loginStatus = login(url)
    try:
        time.sleep(1)
        driver.find_element_by_xpath(jz_manage).click()  # 点击基站管理
        time.sleep(1)

        # set_output_40dbm(driver)
        printf("点击功放管理")
        driver.find_element_by_xpath(pa_manage).click()
        time.sleep(3)

        driver.find_element_by_xpath(refresh).click()
        printf("点击刷新")
        time.sleep(3)

        paRetStatus[indexNu] = get_current_time_str()
        indexNu += 1

        time.sleep(2)
        driver.find_element_by_xpath(pa_off).click()
        printf("关闭所有功放：")
        printf(datetime.now())
        time.sleep(5)

        # add 当功放开关操作出错警告是，出现弹框，
        pa_warning_detect(driver)

        driver.find_element_by_xpath(refresh).click()
        printf("点击刷新")
        time.sleep(5)

        # 读取表内容，循环获取三个功放值
        cssStrPa1 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(1) > td.el-table_1_column_9 > div > div > span'
        cssStrPa2 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(2) > td.el-table_1_column_9 > div > div > span'
        cssStrPa3 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(3) > td.el-table_1_column_9 > div > div > span'
        cssClosePaStatus1 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(1) > td.el-table_1_column_12 > div > div > span'
        cssClosePaStatus2 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(2) > td.el-table_1_column_12 > div > div > span'
        cssClosePaStatus3 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(3) > td.el-table_1_column_12 > div > div > span'
        xpath_option_values_on1 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[1]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(19, 206, 102); background-color: rgb(19, 206, 102);"]'
        xpath_option_values_on2 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[2]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(19, 206, 102); background-color: rgb(19, 206, 102);"]'
        xpath_option_values_on3 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[3]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(19, 206, 102); background-color: rgb(19, 206, 102);"]'
        xpath_option_values_off1 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[1]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(255, 73, 73); background-color: rgb(255, 73, 73);"]'
        xpath_option_values_off2 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[2]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(255, 73, 73); background-color: rgb(255, 73, 73);"]'
        xpath_option_values_off3 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[3]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(255, 73, 73); background-color: rgb(255, 73, 73);"]'
        strIn = [""] * 4
        strIn[1] = cssStrPa1
        strIn[2] = cssStrPa2
        strIn[3] = cssStrPa3
        ccpsIn = [""] * 4
        ccpsIn[1] = cssClosePaStatus1
        ccpsIn[2] = cssClosePaStatus2
        ccpsIn[3] = cssClosePaStatus3
        xovon = [""] * 4
        xovon[1] = xpath_option_values_on1
        xovon[2] = xpath_option_values_on2
        xovon[3] = xpath_option_values_on3
        xovoff = [""] * 4
        xovoff[1] = xpath_option_values_off1
        xovoff[2] = xpath_option_values_off2
        xovoff[3] = xpath_option_values_off3

        for i in {1, 2, 3}:
            try:
                driver, ret_val = get_pa_val(driver, strIn[i])  # 获取功放值
                paCheckValus[i - 1] = ret_val

                tar = ret_val.index('dBm')
                ret_val1 = ret_val[:tar]
                val = float(ret_val1)
                if val == -999.0 or val > 0.0:
                    printf("关闭失败")
                    paRetStatus[indexNu] = "false"
                    indexNu += 1
                elif val > -50.0 and val <= 0.0:
                    printf("功放关闭成功")
                    paRetStatus[indexNu] = "pass"
                    indexNu += 1
            except Exception as err:
                printf(err)
                paRetStatus[indexNu] = "false"
                indexNu += 1
                printf(sys._getframe().f_lineno)
                continue

        time.sleep(2)
        # 点击开启功放,3秒后，点击刷新，读取功放输出功率为（38-40）dbm
        indexNu = 4
        paRetStatus[indexNu] = get_current_time_str()
        indexNu += 1
        driver.find_element_by_xpath(pa_on).click()
        printf("开启所有功放")
        time.sleep(5)

        # add 当功放开关操作出错警告是，出现弹框，
        pa_warning_detect(driver)

        driver.find_element_by_xpath(refresh).click()
        printf("点击刷新")
        time.sleep(5)
        for i in {1, 2, 3}:  # 循环获取三个功放值
            try:
                driver, ret_val = get_pa_val(driver, strIn[i])  # 获取功放值
                paCheckValus[i - 1 + 3] = ret_val
                tar = ret_val.index('dBm')
                ret_val1 = ret_val[:tar]
                val = float(ret_val1)
                if val == -999.0:
                    printf("读取功放失败")
                    paRetStatus[indexNu] = "false"
                    indexNu += 1
                elif val > -50.0 and val <= 0.0:
                    printf("功放打开失败")
                    paRetStatus[indexNu] = "false"
                    indexNu += 1
                elif val > 0.0:
                    printf("功放打开成功")
                    paRetStatus[indexNu] = "pass"
                    indexNu += 1
            except Exception as err:
                printf(err)
                paRetStatus[indexNu] = "false"
                indexNu += 1
                printf(sys._getframe().f_lineno)
                continue
        driver.quit()
        return paRetStatus, paCheckValus, 1
    except Exception as err:
        printf(err)
        printf(sys._getframe().f_lineno)
        driver.quit()
        return paRetStatus, paCheckValus, -1


def PA_Reboot_opera(url, status):
    # 读取表内容，循环获取三个功放值
    cssStrPa1 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(1) > td.el-table_1_column_9 > div > div > span'
    cssStrPa2 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(2) > td.el-table_1_column_9 > div > div > span'
    cssStrPa3 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(3) > td.el-table_1_column_9 > div > div > span'
    cssClosePaStatus1 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(1) > td.el-table_1_column_12 > div > div > span'
    cssClosePaStatus2 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(2) > td.el-table_1_column_12 > div > div > span'
    cssClosePaStatus3 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr:nth-child(3) > td.el-table_1_column_12 > div > div > span'
    xpath_option_values_on1 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[1]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(19, 206, 102); background-color: rgb(19, 206, 102);"]'
    xpath_option_values_on2 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[2]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(19, 206, 102); background-color: rgb(19, 206, 102);"]'
    xpath_option_values_on3 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[3]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(19, 206, 102); background-color: rgb(19, 206, 102);"]'
    xpath_option_values_off1 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[1]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(255, 73, 73); background-color: rgb(255, 73, 73);"]'
    xpath_option_values_off2 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[2]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(255, 73, 73); background-color: rgb(255, 73, 73);"]'
    xpath_option_values_off3 = '//div[@class="el-table__body-wrapper is-scrolling-left"]/table/tbody/tr[3]/td[12]/div/div/span[@style="width: 40px; border-color: rgb(255, 73, 73); background-color: rgb(255, 73, 73);"]'
    strIn = [""] * 4
    strIn[1] = cssStrPa1
    strIn[2] = cssStrPa2
    strIn[3] = cssStrPa3
    ccpsIn = [""] * 4
    ccpsIn[1] = cssClosePaStatus1
    ccpsIn[2] = cssClosePaStatus2
    ccpsIn[3] = cssClosePaStatus3
    xovon = [""] * 4
    xovon[1] = xpath_option_values_on1
    xovon[2] = xpath_option_values_on2
    xovon[3] = xpath_option_values_on3
    xovoff = [""] * 4
    xovoff[1] = xpath_option_values_off1
    xovoff[2] = xpath_option_values_off2
    xovoff[3] = xpath_option_values_off3

    driver, loginStatus = login(url)
    if loginStatus == 1:
        printf("登录成功啦")
    else:
        printf("登录失败，等待10s后再次尝试登录")
        time.sleep(10)
        driver, loginStatus = login(url)
    try:
        time.sleep(1)
        driver.find_element_by_xpath(jz_manage).click()  # 点击基站管理
        time.sleep(1)

        # set_output_40dbm(driver)
        printf("点击功放管理")
        driver.find_element_by_xpath(pa_manage).click()
        time.sleep(3)

        driver.find_element_by_xpath(refresh).click()
        printf("点击刷新")
        time.sleep(3)

        if status == False:  # 关闭功放
            time.sleep(2)
            driver.find_element_by_xpath(pa_off).click()
            printf("关闭所有功放：")
            printf(datetime.now())
            time.sleep(5)
            driver.find_element_by_xpath(refresh).click()
            driver, pp_ret_val = get_pa_property(driver, xovoff[1])
            if pp_ret_val == 0:
                pass
            else:
                pass
            driver, ret_val = get_pa_val(driver, strIn[1])  # 获取功放值
            val = int(ret_val)
            if val == -999 or val > 0:
                printf("关闭失败")
            elif val > -50 and val <= 0 and pp_ret_val == 0:
                printf("功放关闭成功")

        else:  # 开启功放
            driver.find_element_by_xpath(pa_on).click()
            printf("开启所有功放")
            time.sleep(5)
            driver.find_element_by_xpath(refresh).click()
            printf("点击刷新")
            time.sleep(5)

            driver, pp_ret_val = get_pa_property(driver, xovon[1])
            if pp_ret_val == 0:
                pass
            else:
                pass
            driver, ret_val = get_pa_val(driver, strIn[1])  # 获取功放值
            val = int(ret_val)
            if val == -999:
                printf("读取功放失败")

            elif val > -50 and val <= 0:
                printf("功放打开失败")

            elif val > 0 and pp_ret_val == 0:
                printf("功放打开成功")

        driver.quit()
        return 1
    except Exception as err:
        printf(err)
        printf(sys._getframe().f_lineno)
        driver.quit()
        return -1


'''
函数名：KB_Reboot
功能：
    核心板软重启，核心板断电重启
流程：
    初始化参数
    登录
    击基站管理 jz_manage
    点击核心板管理
    点击第band_n块板
    点击rebootType重启
    点击重启确认按钮

入参：
    url:地址
    rebootType：重启类型
    band_n：选择第几块板
返回值：
    ['重启时间','建立小区时间'结果']
    err_flags：
        ==0:成功
        < 0:失败
修改时间：
    2020年4月29日13:43:23
     2020年5月5日12:06:49
        添加错误标志返回值err_flags
'''


def KB_Reboot(url, rebootType, band_n):
    kbRetStatus = -1
    err_flags = 0
    kbStatus = ['', '', 'false']
    main_xpath = '//*[@id="app"]/section/section/main'
    fresh_click = 'main > div > div.el-tabs__content > div.box-card > div.search.el-row > div > button > span'
    first_css_v0429 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr.el-table__row.expanded > td.el-table_1_column_7 > div > div > span'
    css_cell_build_status = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr.el-table__row.expanded > td.el-table_1_column_9 > div > div > span'
    driver, loginStatus = login(url)
    if loginStatus == 1:
        printf("登录成功啦")
    else:
        printf("登录失败，等待10s后再次尝试登录")
        time.sleep(10)
        driver, loginStatus = login(url)

    kbStatus[0] = get_current_time_str()
    try:
        time.sleep(1)
        driver.find_element_by_xpath(jz_manage).click()
        printf("点击基站管理")
        time.sleep(1)

        driver.find_element_by_xpath(band_manage).click()
        printf("点击核心板管理")
        time.sleep(1)
        ret = click_band_n(driver, band_n)
        if ret == -1:
            printf("点击失败，可能该板子不存在")
            err_flags = -2
            driver.quit()
            return kbStatus, err_flags

        printf("点击第band_n块板:")
        printf(band_n)
        time.sleep(1)

        driver.find_element_by_xpath(rebootType).click()
        printf("点击rebootType重启：")
        printf(rebootType)
        time.sleep(1)

        driver.find_element_by_xpath(comfirm).click()
        printf("点击重启确认")
        time.sleep(1)
        # 第三次循环开始，每分钟查询一次状态，直到成功状态就退出，否则直接等待待最后结束判断结果
        for i in {1, 2, 3, 4, 5, 6, 7}:
            if i == 1:
                printf("等待5s")
                time.sleep(5)
            elif i == 2:
                printf("等待20s")
                time.sleep(20)
            else:
                printf("等待60s")
                for j in range(3):
                    time.sleep(18)
                    locater = driver.find_element_by_xpath(main_xpath)
                    time.sleep(2)
                    locater.find_element_by_css_selector(fresh_click).click()
                    printf("点击刷新")

            time.sleep(1)
            locater = driver.find_element_by_xpath(main_xpath)
            time.sleep(1)
            locater.find_element_by_css_selector(fresh_click).click()
            printf("点击刷新")
            time.sleep(1)
            locater = driver.find_element_by_xpath(main_xpath)
            retStr = locater.find_element_by_css_selector(first_css_v0429).text

            if "离线" in retStr:
                printf("离线状态")
            elif "在线" in retStr:
                retCellStatus = locater.find_element_by_css_selector(css_cell_build_status).text
                printf("小区建立状态查询：" + retCellStatus)
                if "未知" in retCellStatus:
                    pass
                elif "取消建立中" in retCellStatus:
                    pass
                elif "小区建立中" in retCellStatus:
                    pass
                elif "未建立" in retCellStatus:
                    pass
                elif "建立成功" in retCellStatus:
                    kbRetStatus += 1
                    if kbStatus[1] == '':
                        kbStatus[1] = get_current_time_str()
                    # 在线状态并且建立成功，就结束退出
                    break
                elif "建立失败" in retCellStatus:
                    pass
                else:
                    pass
            else:
                printf("非离线-在线状态：")
        time.sleep(1)
        printf("关闭浏览器")

    except Exception as err:
        printf(err)
        printf(sys._getframe().f_lineno)
        printf("无法连接，请检查设备的是否开启，或是否在同一网段")
        err_flags = -1
    driver.quit()
    if kbRetStatus >= 0:  # 重启成功
        kbStatus[2] = "pass"
    else:
        kbStatus[2] = "false"
    return kbStatus, err_flags


#     添加温度核心板温度查询
# 实时数据-核心板温度-温度数据获取
def kb_temperature_check(url):
    result1 = []
    result2 = []
    driver, loginStatus = login(url)

    driver.find_element_by_xpath(realtimedata).click()
    printf("点击实时数据")
    time.sleep(1)

    driver.find_element_by_xpath(kb_temperature).click()
    printf("点击核心板温度")
    time.sleep(1)

    fapName = '//span[@class="fapName"]'
    res = driver.find_elements_by_xpath(fapName)
    for i in res:
        printf(i.text)
        result1.append(i.text)

    tagTems = '//span[@class="tagTem el-tag el-tag--success el-tag--dark"]'
    res = driver.find_elements_by_xpath(tagTems)
    for i in res:
        printf(i.text)
        result2.append(i.text)
    driver.quit()

    result3 = []
    for v in range(len(res)):
        result3.append(result1[v])
        result3.append(result2[v])

    return result3

def KB_Reboot_Step1(url, rebootType):
    # 0 1 2
    # 3 4 5
    # 6 7 8
    kbStatus = ['', '', 'false',
                '', '', 'false',
                '', '', 'false',
                ]
    buildResult = False

    main_xpath = '//*[@id="app"]/section/section/main'
    fresh_click = 'main > div > div.el-tabs__content > div.box-card > div.search.el-row > div > button > span'
    first_css_v0429 = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr.el-table__row.expanded > td.el-table_1_column_7 > div > div > span'
    css_cell_build_status = 'main > div > div.el-tabs__content > div.box-card > div.el-table.el-table--fit.el-table--border.el-table--scrollable-x.el-table--enable-row-hover.el-table--enable-row-transition.el-table--mini > div.el-table__body-wrapper.is-scrolling-left > table > tbody > tr.el-table__row.expanded > td.el-table_1_column_9 > div > div > span'
    driver, loginStatus = login(url)
    if loginStatus == 1:
        printf("登录成功啦")
    else:
        printf("登录失败，等待10s后再次尝试登录")
        time.sleep(10)
        driver, loginStatus = login(url)

    try:
        time.sleep(1)
        driver.find_element_by_xpath(jz_manage).click()
        printf("点击基站管理")
        time.sleep(1)

        driver.find_element_by_xpath(band_manage).click()
        printf("点击核心板管理")
        time.sleep(1)

        # 循环三次点击板子重启
        bandNum = 3
        for i in {1, 2, 3}:
            band = str.format("//div[text()='%d']" % i)
            ret = click_band_n(driver, band)
            if ret == -1:
                printf("点击失败，可能该板子不存在")
                bandNum -= 1
                continue

            driver.find_element_by_xpath(rebootType).click()
            printf("点击rebootType重启：")
            printf(rebootType)
            time.sleep(1)

            driver.find_element_by_xpath(comfirm).click()
            printf("点击重启确认")
            time.sleep(1)

            kbStatus[i * 3 - 3] = get_current_time_str()

        # 第三次循环开始，每分钟查询一次状态，直到成功状态就退出，否则直接等待待最后结束判断结果
        for i in {1, 2, 3, 4, 5, 6}:
            if buildResult:
                break

            if i == 1:
                printf("等待5s")
                time.sleep(5)
            elif i == 2:
                printf("等待20s")
                time.sleep(20)
            else:
                printf("等待60s")
                for j in range(3):
                    time.sleep(18)
                    locater = driver.find_element_by_xpath(main_xpath)
                    time.sleep(2)
                    locater.find_element_by_css_selector(fresh_click).click()
                    printf("点击刷新")

            for i in range(bandNum):
                n = i + 1
                band = str.format("//div[text()='%d']" % n)
                ret = click_band_n(driver, band)
                if ret == -1:
                    printf("点击失败，可能该板子不存在")
                    err_flags = -2
                    driver.quit()
                    return kbStatus, err_flags

                time.sleep(1)
                locater = driver.find_element_by_xpath(main_xpath)
                time.sleep(1)
                locater.find_element_by_css_selector(fresh_click).click()
                printf("点击刷新")
                time.sleep(1)
                locater = driver.find_element_by_xpath(main_xpath)
                retStr = locater.find_element_by_css_selector(first_css_v0429).text

                if "离线" in retStr:
                    printf("离线状态")
                elif "在线" in retStr:
                    retCellStatus = locater.find_element_by_css_selector(css_cell_build_status).text
                    printf("小区建立状态查询：" + retCellStatus)
                    if "未知" in retCellStatus:
                        pass
                    elif "取消建立中" in retCellStatus:
                        pass
                    elif "小区建立中" in retCellStatus:
                        pass
                    elif "未建立" in retCellStatus:
                        pass
                    elif "建立成功" in retCellStatus:
                        buildstatus = 0
                        # 147
                        index1 = i * 3 + 1
                        # 258
                        index2 = i * 3 + 2
                        kbStatus[index2] = 'pass'
                        kbStatus[index1] = get_current_time_str()
                        for j in range(bandNum):
                            index3 = j * 3 + 2
                            if kbStatus[index3] == 'pass':
                                buildstatus += 1

                        if buildstatus == bandNum:
                            printf('多块板均建立小区成功，退出')
                            buildResult = True
                            break

                    elif "建立失败" in retCellStatus:
                        pass
                    else:
                        pass
                else:
                    printf("非离线-在线状态：")

        time.sleep(1)
        printf("关闭浏览器")

    except Exception as err:
        printf(err)
        printf(sys._getframe().f_lineno)
        printf("无法连接，请检查设备的是否开启，或是否在同一网段")
        driver.quit()
        return kbStatus, buildResult

    driver.quit()
    return kbStatus, buildResult


'''
函数名：gpio_power_down_reboot
功能：
    执行gpio多块板的重启操作
流程：
    初始化参数
    登录
    击基站管理 jz_manage
    点击核心板管理
    点击第band_n块板
    点击rebootType重启
    点击重启确认按钮

入参：
    url:地址
    rebootType：重启类型
    band_n：选择第几块板
返回值：
    ['重启时间','建立小区时间'结果']
    err_flags：
        ==0:成功
        < 0:失败
修改时间：
    2020年4月29日13:43:23
     2020年5月5日12:06:49
        添加错误标志返回值err_flags
'''


def gpio_power_down_reboot(url, band_n):
    driver, loginStatus = login(url)
    if loginStatus == 1:
        printf("登录成功")
    else:
        printf("登录失败，等待10s后再次尝试登录")
        time.sleep(10)
        driver, loginStatus = login(url)

    try:
        time.sleep(1)
        driver.find_element_by_xpath(jz_manage).click()
        printf("点击基站管理")
        time.sleep(1)

        driver.find_element_by_xpath(band_manage).click()
        printf("点击核心板管理")
        time.sleep(1)
        click_band_n(driver, band_n)
        time.sleep(1)
        driver.find_element_by_xpath(power_outage_restart).click()
        time.sleep(1)
        driver.find_element_by_xpath(comfirm_button).click()
        printf("点击重启确认")
        timeStr = get_current_time_str()
        time.sleep(1)
        driver.quit()
        return timeStr
    except Exception as err:
        printf(err)
        printf(sys._getframe().f_lineno)
        driver.quit()
        return ""


'''
函数名：MCB_Reboot
功能：
    主控板重启测试；（每次重启的间隔时间为30min）；
    标准：
    A、点击重启主控板后，返回主控重启成功的字段；
    B、5秒后登陆控制台（输入映射的IP），登陆失败；
    C、1min钟后登陆控制台（输入映射的IP），登陆成功，能查询到主控版本信息；
    上述三点均满足，则判断此次主控板重启成功；
参数：
    url:地址
返回值：
    ['重启时间','建立小区时间',结果']
    mcb_flags：
        ==0:成功
        < 0:失败
修改时间：
    2020年4月29日13:40:15
    2020年4月30日14:03:02
        修改点击重启后延时时间为10s,原本5s时间过短
    2020年5月5日12:06:49
        添加错误标志返回值mcb_flags
'''


def MCB_Reboot(url):
    printf("打开浏览器，登录")
    mcb_flags = 0
    mcbRetStatus = ['', '', 'false']
    driver, loginStatus = login(url)
    if loginStatus == 1:
        printf("登录成功")
    else:
        printf("登录失败，等待10s后再次尝试登录")
        time.sleep(10)
        driver, loginStatus = login(url)
        if loginStatus == 1:
            printf("登录成功")
        else:
            return mcbRetStatus, -1

    mcbRetStatus[0] = get_current_time_str()
    try:
        driver.find_element_by_xpath(reboot_mcband).click()
        time.sleep(1)
        driver.find_element_by_xpath(comfirm).click()
        time.sleep(1)
        printf("关闭浏览器")
        driver.quit()
        printf('等待10s，登录控制台：预计结果为登录失败')
        time.sleep(10)
        loginStatus = login_test(url)
        if loginStatus == 0:
            printf("登录成功,主控为进行重启，重启失败")
            driver.quit()
            return mcbRetStatus, -1
        else:
            printf("登录失败,主控正在重启中，请等待")
            driver.quit()

        printf("等待一分钟后登录控制台，预计结果为登录成功")
        time.sleep(60)
        for i in {1, 2}:
            loginStatus = login_test(url)
            if loginStatus == 0:
                printf("重启主控成功")
                mcbRetStatus[1] = get_current_time_str()
                mcbRetStatus[2] = 'pass'
                break
            else:
                if i == 1:
                    printf("重启主控失败,等待30s后再尝试")
                    time.sleep(30)
                    continue
                else:
                    printf("重启主控失败")
                    mcb_flags = -1
                    break

        printf("退出浏览器")
    except Exception as err:
        printf(err)
        mcb_flags = -1
    driver.quit()
    return mcbRetStatus, mcb_flags


################应用接口下##########################

################底层接口上##########################
def get_current_time_str():
    update_str = time.strftime('%Y-%m-%d %H:%M:%S')
    return update_str


'''
函数名：printf
功能：打印日志，控制台输出
参数：输出文本
返回值：无
修改时间：
    2020年4月29日14:39:44
    2020年5月5日14:21:29
        将进程号与ip对应，通过进程号获取到ip,然后打印
'''


def printf(test):
    ip = pid_ip_dictionary.get(threading.currentThread().ident)
    logger.info("[%s]%s" % (ip, test))


'''
函数名：login
功能：
    登录
入参：
    url:地址
返回值：
    driver：登录页面句柄
    loginStatus：true:登录成功， false:登录失败
修改时间：
    2020年4月29日14:38:01
    2020年5月4日10:41:42
        添加返回值
'''


def login_v2(url,func,isRun):
    printf("打开浏览器")

    # 打开浏览器
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')

    # 创建浏览器对象
    driver = webdriver.Chrome(executable_path=Chrome_path, chrome_options=chrome_options)

    # driver = webdriver.Chrome(Chrome_path)
    driver.implicitly_wait(10)
    try:
        driver.get(url)
        driver.maximize_window()
        printf("输入用户名：" + LOGIN_USER)
        driver.find_element_by_xpath(inputuser_box).send_keys(LOGIN_USER)
        printf("输入密码：" + LOGIN_PASSWD)
        driver.find_element_by_xpath(inputpasswd_box).send_keys(LOGIN_PASSWD)
        driver.find_element_by_xpath(login_submit_btn).click()
        printf("点击了登录按钮")
        time.sleep(2)
        if isRun:
            id = func(driver)
            printf("dev Id:" + id)
        return driver, 1
    except Exception as err:
        printf(err)
        printf(sys._getframe().f_lineno)
        driver.quit()
        return driver, 0

def login(url):
    printf("打开浏览器")

    # 打开浏览器
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')

    # 创建浏览器对象
    driver = webdriver.Chrome(executable_path=Chrome_path, chrome_options=chrome_options)

    # driver = webdriver.Chrome(Chrome_path)
    driver.implicitly_wait(10)
    try:
        driver.get(url)
        driver.maximize_window()
        printf("输入用户名：" + LOGIN_USER)
        driver.find_element_by_xpath(inputuser_box).send_keys(LOGIN_USER)
        printf("输入密码：" + LOGIN_PASSWD)
        driver.find_element_by_xpath(inputpasswd_box).send_keys(LOGIN_PASSWD)
        driver.find_element_by_xpath(login_submit_btn).click()
        printf("点击了登录按钮")
        time.sleep(2)
        id = find_dev_id(driver)
        printf("dev Id:" + id)
        return driver, 1
    except Exception as err:
        printf(err)
        printf(sys._getframe().f_lineno)
        driver.quit()
        return driver, 0


'''
函数名：login_test
功能：
    登录
入参：
    url:地址
返回值：
    0：登录成功
    -1:登录失败
修改时间：2020年4月29日14:38:01
'''


def login_test(url):
    printf("打开浏览器")
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')

    # 创建浏览器对象
    driver = webdriver.Chrome(executable_path=Chrome_path, chrome_options=chrome_options)
    # driver = webdriver.Chrome(Chrome_path)
    driver.implicitly_wait(10)
    try:
        driver.get(url)
        driver.maximize_window()
        printf("开始登录操作")
        driver.find_element_by_xpath(inputuser_box).send_keys(LOGIN_USER)
        driver.find_element_by_xpath(inputpasswd_box).send_keys(LOGIN_PASSWD)
        driver.find_element_by_xpath(login_submit_btn).click()
        printf("点击了登录按钮")

    except:
        printf("登录异常,关闭浏览器")
        driver.quit()
        printf(sys._getframe().f_lineno)
        return -1
    else:
        printf("登录成功")
        time.sleep(1)
        driver.quit()
        return 0


'''
函数名：login_test_v2
功能：
    登录
入参：
    url:地址
返回值：
    0：登录成功
    -1:登录失败
修改时间：
    2020年4月29日14:38:01
'''


def login_test_v2(url, nu):
    time.sleep(2)
    printf("打开浏览器")
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')

    # 创建浏览器对象
    driver = webdriver.Chrome(executable_path=Chrome_path, chrome_options=chrome_options)
    # driver = webdriver.Chrome(Chrome_path)
    driver.implicitly_wait(10)
    try:
        driver.get(url)
        driver.maximize_window()
        printf("开始登录操作")
        driver.find_element_by_xpath(inputuser_box).send_keys(LOGIN_USER)
        driver.find_element_by_xpath(inputpasswd_box).send_keys(LOGIN_PASSWD)
        driver.find_element_by_xpath(login_submit_btn).click()
        printf("点击了登录按钮")
        time.sleep(2)
        id = find_dev_id(driver)
        print("dev Id:%s" % id)
        lock.acquire()
        pid_div_id_dictionary.setdefault(threading.currentThread().ident, id)
        lock.release()
    except:
        printf("登录异常,关闭浏览器")
        driver.quit()
        printf(sys._getframe().f_lineno)
        return -1
    else:
        driver.quit()
        printf("登录成功")
        return 0


'''
函数名：login_get_device_id
功能：
    登录
入参：
    url:地址
返回值：
    0：登录成功
    -1:登录失败
修改时间：
    2020年4月29日14:38:01
'''


def login_get_device_id(url):
    time.sleep(2)
    printf("打开浏览器")
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')

    # 创建浏览器对象
    driver = webdriver.Chrome(executable_path=Chrome_path, chrome_options=chrome_options)
    # driver = webdriver.Chrome(Chrome_path)
    driver.implicitly_wait(10)
    id = 0
    try:
        driver.get(url)
        driver.maximize_window()
        printf("开始登录操作")
        driver.find_element_by_xpath(inputuser_box).send_keys(LOGIN_USER)
        driver.find_element_by_xpath(inputpasswd_box).send_keys(LOGIN_PASSWD)
        driver.find_element_by_xpath(login_submit_btn).click()
        printf("点击了登录按钮")
        # 等待18秒
        for i in range(3):
            printf("等待10秒")
            time.sleep(10)

        id = find_dev_id(driver)
        print("dev Id:%s" % id)

    except:
        printf("登录异常,关闭浏览器")
        driver.quit()
        printf(sys._getframe().f_lineno)
        return id, -1
    else:
        driver.quit()
        printf("登录成功")
        return id, 0


def find_dev_id(driver):
    devIdxpath = '//*[@id="app"]/section/header/div[1]/h2'
    s = ""
    try:
        id = driver.find_element_by_xpath(devIdxpath).text
        pre = id.index(':') + 1
        tar = id.index(']')
        s = id[pre:tar]
    except Exception as err:
        printf("定位设备id失败，s:%s" % s)
    return s


'''
函数名：set_output_40dbm
功能:设置核心板输出功率为40dbm
返回值:无
修改时间：2020年4月29日14:42:25
'''


def set_output_40dbm(driver):
    # 点击核心板管理
    driver.find_element_by_xpath(band_manage).click()
    # time.sleep(1)
    # 点击第一块板子
    click_band_1(driver)
    # 点击参数设置
    click_setpar_band(driver)
    # 设置输出功率为40dbm
    type_UePMax(driver, "40")
    time.sleep(1)
    # 保存
    click_save_par(driver)
    time.sleep(1)
    click_save_par_confirm(driver)
    time.sleep(1)


'''
函数名：get_pa_val
功能：读取功放值
入参：
    driver:页面句柄
    cssStr：css元素定位
返回值：
    driver：页面句柄
    ret_val：功放输出功率值
修改时间：2020年4月29日14:45:15
'''


def get_pa_val(driver, cssStr):
    ret_val = -999.0
    try:
        main_xpath = '//main[@class="el-main main"]'
        locater = driver.find_element_by_xpath(main_xpath)
        ret_val = locater.find_element_by_css_selector(cssStr).text
        printf('功放输出功率为：' + ret_val)
    except Exception as err:
        printf(err)
    return driver, ret_val


'''
函数名：get_pa_property
功能：读取按钮属性值
入参：
    driver:页面句柄
    cssStr：css元素定位
返回值：
    driver：页面句柄
    ret_val：功放输出功率值
修改时间：2020年4月29日14:45:15
'''


def get_pa_property(driver, xpathStr):
    flags = 0
    try:
        main_xpath = '//main[@class="el-main main"]'
        locater = driver.find_element_by_xpath(main_xpath)
        locater.find_element_by_xpath(xpathStr)
        printf("定位属性成功")
    except Exception as err:
        printf("定位属性失败")
        printf(err)
        flags = -1
    return driver, flags


# 点击第一块基带板
def click_band_1(driver):
    printf("点击第一块基带板")
    driver.find_element_by_xpath(band_1).click()
    time.sleep(1)


'''
函数名：click_band_n
功能：点击第band_n块包板
入参：
    driver：句柄
    band_n：第n块板xpath
返回值：
修改时间：2020年4月29日14:47:05
'''


def click_band_n(driver, band_n):
    try:
        printf(band_n)
        driver.find_element_by_xpath(band_n).click()
        time.sleep(1)
        return 1
    except Exception as err:
        printf(err)
        return -1


# 点击参数设置
def click_setpar_band(driver):
    # self.clicks(self.setpar_band)
    printf("点击参数设置")
    driver.find_element_by_xpath(driver.setpar_band).click()
    time.sleep(1)


# 清空并输入输出功率(dBm) clear() send_keys
def type_UePMax(driver, text):
    printf("清空并输入输出功率")
    driver.find_element_by_xpath(UePMax).clear()
    driver.find_element_by_xpath(UePMax).send_keys(text)


# 保存参数
def click_save_par(driver):
    printf("保存参数")
    driver.find_element_by_xpath(save_par).click()


# 保存参数成功，点击确定按钮
def click_save_par_confirm(driver):
    printf("点击确定按钮")
    driver.find_element_by_xpath(comfirm_button).click()

class Thread_1(QThread):  # 线程1
    trigger = pyqtSignal()
    counter = 0
    def __init__(self):
        super().__init__()
    def run(self):
        printf(self.counter)
        while True:
            if self.counter <= 0:
                break

            time.sleep(1)
            self.trigger.emit()
            self.counter -= 1


##############底层接口下############################

class MyWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MyWindow, self).__init__(parent)
        self.setupUi(self)
        '''
        界面初始化，设置槽函数
        界面默认参数设置
        读取配置文件，获取ip组
        表格初始化，显示更新ip组
        参数界面参数获取传递
        创建线程
        界面线程与执行线程代码标志位响应
        '''
        self.ui_connect_init()
        self.ui_defaule_elemts_init()
        self.ui_read_config_init()
        self.ui_table_init()

    def __del__(self):
        print("End.")

    '''
       函数名：ping_ip
       功能：
           ping ip
           成功则录入字典
       参数：无
       返回值：无
       修改时间：
       2020年5月11日09:51:13

       '''

    def ping_ip(self, ip):
        print("ping ip:%s" % ip)
        backInfo = os.system(u'ping -n 1 -w 1 %s' % ip)
        printf(backInfo)
        if backInfo == 0:
            ip_online_dictionary.update({ip: "pass"})
            printf("pass")
        else:
            ip_online_dictionary.update({ip: "false"})
            printf("false")

    def ping_ip_2(self, ip):
        print("ping ip:%s" % ip)
        backInfo = os.system(u'ping -n 1 -w 1 %s' % ip)
        if backInfo == 0:
            ip_online_dictionary.update({ip: "pass"})
            printf("pass")
            return True
        else:
            ip_online_dictionary.update({ip: "false"})
            printf("false")
            return False

    '''
     函数名：ip_filter_v3
     功能：在线设备过滤
     参数：无
     返回值：无
     修改时间：
         2020年5月6日14:13:41
         2020年5月6日15:27:17
            使用全局字典保存测试过滤的结果
            ip_online_dictionary = {}
    '''
    g_online_device = [""]  # 记录在线ip

    def ip_filter_v3(self):
        # 从表格中提取ip,并且执行ping测试
        for i in range(10):  # 修改为10行，原本为5
            for j in range(12):
                ip = self.tableWidget.item(i, j).text()
                if ip != "":
                    self.ping_ip(ip)  # 结果保存在ip_online_dictionary中

        OnlineIpArry, online_device_num = self.show_table_online_ip_v2()  # 通过全局g_ip_config，ip_online_dictionary判断在线ip
        self.online_device_Edit.setText(str.format("%d" % online_device_num))
        return OnlineIpArry, online_device_num

    '''
       函数名：ip_filter_v4
       功能：在指定的ip中筛选出在线ip
       参数：无
       返回值：无
       修改时间：
           2020年5月6日14:13:41
           2020年5月6日15:27:17
              使用全局字典保存测试过滤的结果
              ip_online_dictionary = {}
      '''

    def ip_filter_v4(self, ipArry):
        online_ip = []
        for ip in ipArry:
            if ip != "":
                if self.ping_ip_2(ip) == True:  # 结果保存在ip_online_dictionary中
                    online_ip.append(ip)
        return online_ip

    '''
    函数名：set_cell_color
    功能：
    参数：无
    返回值：无
    修改时间：
       2020年5月6日14:13:41
       2020年5月6日15:27:17
          使用全局字典保存测试过滤的结果
          ip_online_dictionary = {}
    '''

    def set_cell_color(self, key, color):
        items = self.tableWidget.findItems(key, Qt.MatchExactly)
        if len(items) > 0:
            for i in items:
                i.setBackground(color)  # 设置背景色

    '''
       函数名：show_table_online_ip_v2
       功能：ipArry 通过字典判断是否在线，然后设置背景色即可
       参数：无
       返回值：无
       修改时间：
           2020年5月6日14:13:41
    '''

    def show_table_online_ip_v2(self):
        nu = 0
        OnlineIpArry = [""] * 120  # 60 ->120
        for key in g_ip_config:
            if ip_online_dictionary.get(key) == "pass":
                OnlineIpArry[nu] = key
                nu += 1
                self.set_cell_color(key, online_color)
            else:
                self.set_cell_color(key, offline_color)

        return OnlineIpArry, nu

    '''
         函数名：get_table_online_ip
         功能：获取表格中在线的ip
         参数：无
         返回值：无
         修改时间：
             2020年5月6日14:13:41
    '''

    def get_table_online_ip(self):
        OnlineIpArry = []
        for key in g_ip_config:
            if ip_online_dictionary.get(key) == "pass":
                OnlineIpArry.append(key)
        return OnlineIpArry

    '''
        函数名：MessageBoxDef
        功能：获取表格中在线的ip
        参数：无
        返回值：无
        修改时间：
            2020年5月6日14:13:41
     '''

    def MessageBoxDef(self, title, mesg):
        ret = QMessageBox.question(self, title, mesg, QMessageBox.Yes | QMessageBox.No)
        if ret == QMessageBox.Yes:
            print("确定")
            return True
        elif ret == QMessageBox.No:
            print("取消")
            return False

    '''
        函数名：config_rewrite_data
        功能：
           清空字典
           读取配置文件ip
           录入字典
           记录ip
        参数：无
        返回值：无
        修改时间：
            2020年5月11日09:42:03
    '''

    def config_rewrite_data(self):
        filename = "./config.ini"
        groupName = "IP_Arry"
        keyName = "ip_start"
        curpath = os.path.dirname(os.path.realpath(__file__))
        cfgpath = os.path.join(curpath, "config.ini")

        conf = configparser.ConfigParser()  # 实例化
        conf.read(cfgpath, encoding="utf-8")
        printf(cfgpath)
        conf.set(groupName, keyName, "111111111")
        conf.write(open(filename, "r+", encoding="utf-8"))

    '''
    函数名：config_create_data
    功能：
     清空字典
     读取配置文件ip
     录入字典
     记录ip
    参数：无
    返回值：无
    修改时间：
      2020年5月11日09:42:03
    '''

    def config_create_data(self):
        filename = "config.ini"
        groupName = "IP_Arry"
        conf = configparser.ConfigParser()
        conf.add_section(groupName)
        for i in range(120):  # 60->120
            key = str.format("ip%d" % i)
            index = i + 30
            value = str.format("192.168.3.%d" % index)
            conf.set(groupName, key, value)

        with open(filename, "a+") as f:
            conf.write(f)

    # 创建新excel文件
    '''
    函数名：openpyxl_create_report
    功能：创建新excel文件
    入参：
        fileName：文件名
        tableNames：表名，数组
    返回值：无
    修改时间：2020年4月29日14:54:30
    '''

    def openpyxl_create_report(self, fileName, tableNames):
        path = '../report/' + fileName + ".xlsx"
        data = openpyxl.Workbook()  # 新建工作簿
        i = 0
        for name in tableNames:
            data.create_sheet(name, i)  # 添加页,设置第一个工作页
            i += 1
        data.save(path)  # 一定要保存

    def openpyxl_create_report_new(self, filePath, fileName, tableNames):
        path = filePath + fileName + ".xlsx"
        data = openpyxl.Workbook()  # 新建工作簿
        i = 0
        for name in tableNames:
            print("create table name:%s" % tableNames)
            data.create_sheet(name, i)  # 添加页,设置第一个工作页
            i += 1
        data.save(path)  # 一定要保存
        printf(path)

    def openpyxl_addin_new(self, filePath, fileName, tableName, values, IfHead):
        fileName = filePath + fileName + ".xlsx"
        data = openpyxl.load_workbook(fileName)
        table = data[tableName]
        nrows = table.max_row  # 获得行数+1的值，即下一行应该写入的地方
        if IfHead == 0:
            pass  # 此时为标题头
        else:
            nrows += 1
        col = 1
        resultFlags = 0
        for value in values:
            print("value:%s,row:%d,col:%d" % (value, nrows, col))
            if value is None:
                col += 1
                continue
            table.cell(row=nrows, column=col).value = value
            if 'pass' == value:
                # 设置字体颜色
                table.cell(row=nrows, column=col).fill = green_fill
            elif 'false' == value:
                resultFlags += 1
                table.cell(row=nrows, column=col).fill = red_fill
            # 设置居中对齐
            table.cell(row=nrows, column=col).alignment = Alignment(horizontal='center', vertical='center')
            col += 1
        print(fileName)
        data.save(fileName)

    '''
    函数名：openpyxl_addin
    功能：
        向excel追加内容
        设置字体表格颜色
        设置居中对齐
    入参：
        fileName：文件名
        tableName：表名
        values：写入数组
        IfHead：是否为第一行（0）
    返回值：
    修改时间：
    '''

    def openpyxl_addin(self, fileName, tableName, values, IfHead):
        fileName = '../report/' + fileName + ".xlsx"
        data = openpyxl.load_workbook(fileName)
        # table = data.get_sheet_by_name(tableName)
        table = data[tableName]
        nrows = table.max_row  # 获得行数+1的值，即下一行应该写入的地方
        if IfHead == 0:
            pass  # 此时为标题头
        else:
            nrows += 1
        col = 1
        resultFlags = 0
        for value in values:
            if values == None:
                continue
            print("value:%s,row:%d,col:%d" % (value, nrows, col))
            table.cell(row=nrows, column=col).value = value
            if value == None:
                continue
            if 'pass' in str(value):
                # 设置字体颜色     font
                # 设置单元格颜色   fill
                table.cell(row=nrows, column=col).font = font_pass
            elif 'false' in str(value):
                resultFlags += 1
                table.cell(row=nrows, column=col).font = font_false
            # 设置居中对齐
            table.cell(row=nrows, column=col).alignment = Alignment(horizontal='center', vertical='center')
            # 设置列宽
            col_letter = get_column_letter(col)
            table.column_dimensions[col_letter].width = 20
            col += 1
        data.save(fileName)

    '''
    函数名：thread_runner_pa_mcb_kb_reboot
    功能： 
       功放开关
       核心板软启动
       核心板断电重启
       主控重启
       测试结果输出报告
    参数：
        ip:设备ip
        nu：入参延时
    修改时间：
        2020年4月29日10:43:01
        2020年4月30日14:12:59
            结果记录长度增加到22：原本为values = [''] * 16
            功放无法通过包含值来判断是否为记录结果
    '''

    def thread_runner_pa_mcb_kb_reboot(self, ip, nu):
        global lock  # 声明为全局变量,每个线程函数中都要声明
        time.sleep(1)
        Rep_Num = 0
        values = [''] * 46
        values[Rep_Num] = ip
        Rep_Num += 1

        lock.acquire()
        pid_ip_dictionary.setdefault(threading.currentThread().ident, ip)
        lock.release()

        url = "http://" + ip + "/#/login"
        login_test_v2(url, 1)
        values[Rep_Num] = pid_div_id_dictionary.get(threading.currentThread().ident)
        if values[Rep_Num] == None:
            values[Rep_Num] = ""
        printf(values[1])
        Rep_Num += 1
        print("values[1]:", values[1])
        print("测试次数为：",nu)
        for i in range(nu):
            Rep_Num = 2
            printf(url)

            printf("开始功放开关测试")
            paRetArr, paCheckValus, pa_flags = PA_Reboot(url)
            if pa_flags == -1:
                printf("功放开关测试执行失败,等待10s后再次尝试")
                time.sleep(10)
                paRetArr, paCheckValus, pa_flags = PA_Reboot(url)
                printf("pa_flags:")
                printf(pa_flags)
            else:
                printf("功放开关测试执行成功")

            print("paCheckValus return :%s" % paCheckValus)
            for paValue in paCheckValus:
                values[Rep_Num] = paValue
                Rep_Num += 1

            print("Pa_reboot return :%s" % paRetArr)
            for paValue in paRetArr:
                values[Rep_Num] = paValue
                Rep_Num += 1

            printf("等待1分钟")
            time.sleep(1 * 60)

            kbRetArr, err_flags = KB_Reboot_Step1(url, reboot_band)
            for kbValue in kbRetArr:
                values[Rep_Num] = kbValue
                Rep_Num += 1

            ###断电重启
            printf("等待1分钟")
            time.sleep(1 * 60)
            printf("开始断电重启测试")
            # 核心板断电重启
            kbRetArr, err_flags = KB_Reboot_Step1(url, power_outage_restart)
            for kbValue in kbRetArr:
                values[Rep_Num] = kbValue
                Rep_Num += 1

            # 核心板温度查询
            resTmp = kb_temperature_check(url)
            tmpLen = len(resTmp)
            for kbValue in resTmp:
                values[Rep_Num] = kbValue
                Rep_Num += 1

            if tmpLen >= 0 and tmpLen <= 6:
                Rep_Num += 6 - tmpLen
            else:
                pass

            printf("等待1分钟")
            time.sleep(1 * 60)
            printf("主控重启测试")
            mcbRetArr, mcb_err = MCB_Reboot(url)
            if mcb_err < 0:
                printf("MCB_Reboot出错，等待10s后再次尝试")
                time.sleep(10)
                mcbRetArr, mcb_err = MCB_Reboot(url)
                printf("mcb_err:")
                printf(mcb_err)
            else:
                printf("MCB_Reboot执行成功")
                print("mcbRetArr:%s" % mcbRetArr)
            for mcbValue in mcbRetArr:
                values[Rep_Num] = mcbValue
                Rep_Num += 1

            lock.acquire()  # 上锁，之后的同步代码，只能一个线程访问
            self.openpyxl_addin(filename, "测试结果", values, 1)
            lock.release()  # 解锁
            printf("等待5分钟")
            time.sleep(5 * 60)

        print("END")

    '''
       函数名：thread_runner_mcbsoft_kb_reboot_chaoxun
       功能： 
          功放开关
          核心板软启动
          核心板断电重启
          主控重启
          测试结果输出报告
       参数：
           ip:设备ip
           nu：入参延时
       修改时间：
           只进行基带板软重启，主控重启
       '''

    def thread_runner_mcbsoft_kb_reboot_chaoxun(self, ip, nu):
        global lock  # 声明为全局变量,每个线程函数中都要声明
        time.sleep(1)
        Rep_Num = 0
        values = [''] * 40
        values[Rep_Num] = ip
        Rep_Num += 1

        lock.acquire()
        pid_ip_dictionary.setdefault(threading.currentThread().ident, ip)
        lock.release()

        url = "http://" + ip + "/#/login"
        login_test_v2(url, 1)
        values[Rep_Num] = pid_div_id_dictionary.get(threading.currentThread().ident)
        if values[Rep_Num] == None:
            values[Rep_Num] = ""
        printf(values[1])
        Rep_Num += 1

        for i in range(nu):
            Rep_Num = 2

            kbRetArr, err_flags = KB_Reboot_Step1(url, reboot_band)
            for kbValue in kbRetArr:
                values[Rep_Num] = kbValue
                Rep_Num += 1
            ###主控重启
            printf("等待1分钟")
            time.sleep(1 * 60)
            printf("主控重启测试")
            mcbRetArr, mcb_err = MCB_Reboot(url)
            if mcb_err < 0:
                printf("MCB_Reboot出错，等待10s后再次尝试")
                time.sleep(10)
                mcbRetArr, mcb_err = MCB_Reboot(url)
                printf("mcb_err:")
                printf(mcb_err)
            else:
                printf("MCB_Reboot执行成功")
                print("mcbRetArr:%s" % mcbRetArr)
            for mcbValue in mcbRetArr:
                values[Rep_Num] = mcbValue
                Rep_Num += 1

            lock.acquire()  # 上锁，之后的同步代码，只能一个线程访问
            self.openpyxl_addin(filename, "测试结果", values, 1)
            lock.release()  # 解锁
            printf("等待5分钟")
            time.sleep(5 * 60)

        print("END")

    '''
         函数名：thread_runner_mcbcutdown_kb_reboot_baicaibang
         功能： 
            功放开关
            核心板软启动
            核心板断电重启
            主控重启
            测试结果输出报告
         参数：
             ip:设备ip
             nu：入参延时
         修改时间：
             只进行基带板断电重启，主控重启
         '''

    def thread_runner_mcbcutdown_kb_reboot_baicaibang(self, ip, nu):
        global lock  # 声明为全局变量,每个线程函数中都要声明
        time.sleep(1)
        Rep_Num = 0
        values = [''] * 40
        values[Rep_Num] = ip
        Rep_Num += 1

        lock.acquire()
        pid_ip_dictionary.setdefault(threading.currentThread().ident, ip)
        lock.release()

        url = "http://" + ip + "/#/login"
        login_test_v2(url, 1)
        values[Rep_Num] = pid_div_id_dictionary.get(threading.currentThread().ident)
        if values[Rep_Num] == None:
            values[Rep_Num] = ""
        printf(values[1])
        Rep_Num += 1
        print("values[1]:", values[1])

        for i in range(nu):
            Rep_Num = 2

            kbRetArr, err_flags = KB_Reboot_Step1(url, power_outage_restart)
            for kbValue in kbRetArr:
                values[Rep_Num] = kbValue
                Rep_Num += 1

            ###主控重启
            printf("等待1分钟")
            time.sleep(1 * 60)
            printf("主控重启测试")
            mcbRetArr, mcb_err = MCB_Reboot(url)
            if mcb_err < 0:
                printf("MCB_Reboot出错，等待10s后再次尝试")
                time.sleep(10)
                mcbRetArr, mcb_err = MCB_Reboot(url)
                printf("mcb_err:")
                printf(mcb_err)
            else:
                printf("MCB_Reboot执行成功")
                print("mcbRetArr:%s" % mcbRetArr)
            for mcbValue in mcbRetArr:
                values[Rep_Num] = mcbValue
                Rep_Num += 1

            lock.acquire()  # 上锁，之后的同步代码，只能一个线程访问
            self.openpyxl_addin(filename, "测试结果", values, 1)
            # self.openpyxl_addin_new(g_report_path, filename, "测试结果", values, 1)
            lock.release()  # 解锁
            printf("等待5分钟")
            time.sleep(5 * 60)

        printf("END")

    '''
    函数名：thread_runner_pa_switch
    功能： 
    参数：
        ip:设备ip
        nu：入参延时
    修改时间：
        2020年4月29日10:43:01
        2020年4月30日14:12:59
            结果记录长度增加到22：原本为values = [''] * 16
            功放无法通过包含值来判断是否为记录结果
    '''

    def thread_runner_pa_switch(self, ip, status):
        url = "http://" + ip + "/#/login"
        pa_flags = PA_Reboot_opera(url, status)
        if pa_flags == -1:
            printf("功放开关测试执行失败")
            lock.acquire()
            pa_opera_dict.update({ip: "false"})
            lock.release()
        else:
            printf("功放开关测试执行成功")
            lock.acquire()
            pa_opera_dict.update({ip: "pass"})
            lock.release()

    '''
    函数名：thread_runner_mcb_reboot
    功能： 
    参数：
        ip:设备ip
        nu：入参延时
    修改时间：
        2020年4月29日10:43:01
        2020年4月30日14:12:59
            结果记录长度增加到22：原本为values = [''] * 16
            功放无法通过包含值来判断是否为记录结果
    '''

    def thread_runner_mcb_reboot(self, ip, status):
        url = "http://" + ip + "/#/login"
        # mcbRetStatus = ['ip','id','开始时间', '结束时间', 'false']
        mcbreboot_result = ['ip', 'id', '开始时间', '结束时间', 'false']

        login_test_v2(url, 5)
        lock.acquire()
        pid_ip_dictionary.setdefault(threading.currentThread().ident, ip)
        lock.release()
        mcbreboot_result[0] = ip
        mcbreboot_result[1] = pid_div_id_dictionary.get(threading.currentThread().ident)

        run_timers = int(self.test_num_Edit.text())
        print("执行次数为：%d" % run_timers)
        for i in range(run_timers):
            print("for i value :%d" % i)
            mcbRetStatus, mcb_flags = MCB_Reboot(url)
            mcbreboot_result[2] = mcbRetStatus[0]
            mcbreboot_result[3] = mcbRetStatus[1]
            mcbreboot_result[4] = mcbRetStatus[2]
            print("mcbRetStatus:%s" % mcbRetStatus)
            print("mcb_flags:%d" % mcb_flags)
            if mcb_flags == -1:
                printf("主控重启失败")
            else:
                printf("主控重启成功")

            # 写入报告
            lock.acquire()  # 上锁，之后的同步代码，只能一个线程访问
            self.openpyxl_addin_new(
                self.REPORT_PATH,
                self.MCB_REBOOT_REPORT_FILE_NAME,
                self.MCB_REBOOT_TABLE_NAME[0],
                mcbreboot_result,
                1)
            lock.release()  # 解锁

            if i + 1 == run_timers:
                continue

            printf("等待4分钟")
            time.sleep(4 * 60)

    '''
     函数名：thread_creater_pa_mcb_kb_reboot
     功能：主要业务逻辑调用
     参数：无
     返回值：无
     修改时间：
         2020年5月6日14:13:41
     '''

    def thread_creater_pa_mcb_kb_reboot_testAll(self, IP, nu, function, resultReportHeadLine):
        tableNames = ["测试结果"]
        report_path = self.reportLineEdit.text()
        if os.path.exists(report_path) == False:
            os.mkdir(report_path)  # 目录不存在，创建
        else:
            pass

        try:
            self.openpyxl_create_report_new(report_path, filename, tableNames)
            self.openpyxl_addin_new(report_path, filename, tableNames[0], resultReportHeadLine, 0)
        except Exception as err:
            printf(err)
            return

        threads = []
        for i in range(len(IP)):
            if IP[i] == "":
                continue
            str = "创建线程" + IP[i]
            printf(str)
            thd1 = threading.Thread(target=function, args=(IP[i], nu))  # 创建一个线程
            nu += 1
            threads.append(thd1)

        for th in threads:
            th.start()  # start()---启动线程活动 GDJC6880
            time.sleep(5)
        for th in threads:
            th.join()  # 等待线程结束


    '''
        函数名：thread_creater_mcb_reboot
        功能：主要业务逻辑调用
        参数：无
        返回值：无
        修改时间：
            2020年5月6日14:13:41
    '''
    MCB_REBOOT_TABLE_NAME = ["mcb_reboot", ]
    MCB_REBOOT_REPORT_FILE_NAME = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
    REPORT_PATH = ""

    def thread_creater_mcb_reboot(self, IP, status):
        threads = []
        pa_opera_dict.clear()
        self.MCB_REBOOT_TABLE_NAME = ["mcb_reboot", ]
        self.MCB_REBOOT_REPORT_FILE_NAME = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))

        self.REPORT_PATH = self.reportLineEdit.text()
        if os.path.exists(self.REPORT_PATH) == False:
            os.mkdir(self.REPORT_PATH)  # 目录不存在，创建
        else:
            pass

        try:
            self.openpyxl_create_report_new(
                self.REPORT_PATH,
                self.MCB_REBOOT_REPORT_FILE_NAME,
                self.MCB_REBOOT_TABLE_NAME)

            printf(self.MCB_REBOOT_REPORT_FILE_NAME)
            self.openpyxl_addin_new(
                self.REPORT_PATH,
                self.MCB_REBOOT_REPORT_FILE_NAME,
                self.MCB_REBOOT_TABLE_NAME[0],
                MCB_REBOOT_HEADLINE,
                0)

            printf(self.MCB_REBOOT_REPORT_FILE_NAME)
        except Exception as err:
            print(err)
            return

        for i in range(len(IP)):
            if IP[i] == "":
                continue

            lock.acquire()
            pa_opera_dict.setdefault(IP[i], "false")
            lock.release()

            str = "创建线程" + IP[i]
            printf(str)
            thd1 = threading.Thread(target=self.thread_runner_mcb_reboot, args=(IP[i], status))  # 创建一个线程
            threads.append(thd1)
            time.sleep(2)

        for th in threads:
            th.start()  # start()---启动线程活动 GDJC6880
            time.sleep(5)
        for th in threads:
            th.join()  # 等待线程结束

    '''
       函数名：thread_creater_pa_reboot
       功能：主要业务逻辑调用
       参数：无
       返回值：无
       修改时间：
           2020年5月6日14:13:41
    '''

    def thread_creater_pa_switch(self, IP, status):
        threads = []
        pa_opera_dict.clear()

        for i in range(len(IP)):
            if IP[i] == "":
                continue

            lock.acquire()
            pa_opera_dict.setdefault(IP[i], "false")
            lock.release()

            str = "创建线程" + IP[i]
            printf(str)
            thd1 = threading.Thread(target=self.thread_runner_pa_switch, args=(IP[i], status))  # 创建一个线程
            threads.append(thd1)
            time.sleep(2)

        for th in threads:
            th.start()  # start()---启动线程活动 GDJC6880
            time.sleep(5)
        for th in threads:
            th.join()  # 等待线程结束

    '''
           函数名：ui_read_config_init
           功能：
              清空字典
              读取配置文件ip
              录入字典
              记录ip
           参数：无
           返回值：无
           修改时间：
               2020年5月11日09:42:03
      '''

    def ui_read_config_init(self):
        filename = "../bin/config.ini"
        groupName = "IP_Arry"
        curpath = os.path.dirname(os.path.realpath(__file__))
        cfgpath = os.path.join(curpath, filename)
        conf = configparser.ConfigParser()
        conf.read(cfgpath)
        ip_online_dictionary.clear()  # 清空字典
        for i in range(120):  # 60->120
            keyName = str.format("ip%d" % i)
            ip = conf.get(groupName, keyName)
            if ip == None:
                continue
            ip_online_dictionary.setdefault(ip, "false")  # 初始化，录入字典
            g_ip_config[i] = ip  # 更新ip，记录字典key
        return

    '''
    函数名：ui_table_init
    功能：
        初始化界面表格
        设置行列数
        设置表头隐藏
        表格内容填充
    参数：无
    返回值：无
    修改时间：
        2020年5月6日10:26:32
    '''

    def ui_table_init(self):
        rows = 10  # 修改行数5-》10
        cols = 12
        text_font = QFont("song", 8, QFont.Bold)
        self.tableWidget.setObjectName("table IP")
        self.tableWidget.setColumnCount(cols)
        self.tableWidget.setRowCount(rows)
        self.tableWidget.setCornerButtonEnabled(False)
        self.tableWidget.horizontalHeader().setHighlightSections(False)
        self.tableWidget.verticalHeader().setHighlightSections(False)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 设置表头自适应伸缩模式
        self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

        ip = g_ip_config
        for i in range(rows):
            for j in range(cols):
                self.tableWidget.setItem(i, j, QTableWidgetItem(ip[i * 12 + j]))

        for val in ip:
            items = self.tableWidget.findItems(val, Qt.MatchExactly)
            if len(items) > 0:
                for i in items:
                    i.setBackground(offline_color)  # 设置背景色
                    dict = {val: "false"}
                    ip_online_dictionary.update(dict)

        self.tableWidget.setFont(text_font)
        self.online_device_Edit.setText("0")

    '''
         函数名：ui_connect_init
         功能：界面初始化
            设置表格内容不可修改
            设置界面大小不可拉伸
            设置槽函数
         参数：无
         返回值：无
         修改时间：
             2020年5月6日14:13:41
    '''

    def ui_connect_init(self):
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setFixedSize(self.width(), self.height())

        self.setDriverDirBt.clicked.connect(self.ui_opera_chose_driver_path)
        self.setLogDirBt.clicked.connect(self.ui_opera_chose_log_path)
        self.setRrportDirBt.clicked.connect(self.ui_opera_chose_report_path)

        self.startBt.clicked.connect(self.ui_opera_start)  # start_thread start_test
        self.stopBt.clicked.connect(self.ui_opera_stop)

        self.tableWidget.clicked.connect(self.ui_opera_chose_table_cell)
        self.refreshBt.clicked.connect(self.ui_opera_config_refresh)
        self.SelectAllpushButton.clicked.connect(self.ui_opera_select_all_device)
        # self.pa_pushButton.clicked.connect(self.ui_opera_pa_switch)
        # self.mcbRebootBt.clicked.connect(self.ui_opera_mcb_reboot)
        # self.gpio_power_down_reboot.clicked.connect(self.ui_opera_gpio_power_down_reboot)
        # self.comboBox.currentIndexChanged[str].connect(self.print_value)  # 条目发生改变，发射信号，传递条目内容

    '''
           函数名：ui_defaule_init
           功能：界面默认值初始化
           参数：无
           返回值：无
           修改时间：
               2020年5月6日14:13:41
    '''

    def ui_defaule_elemts_init(self):
        self.logDirLineEdit.setText("../logs/")
        self.reportLineEdit.setText("../report/")
        self.chromeDriverDirLineEdit.setText("../bin/chromedriver.exe")
        # self.ip_filter_RdBt.click()
        self.online_device_Edit.setText("0")
        self.test_num_Edit.setText("30")
        self.tableWidget.setSelectionMode(QAbstractItemView.NoSelection)  # 取消表格鼠标点击选中显示

        #     add 2020年7月10日17:23:15 添加下拉框
        self.comboBox.addItems(['设备ip筛选测试',
                                '热点-大电围',
                                '微热点-百才帮',
                                '微热点-超讯',
                                '主控重启',
                                'GPIO断电重启',
                                '常规测试'])
        # 设置当前的显示字符
        self.comboBox.setCurrentText('设备ip筛选测试')
        # 设置时间显示格式
        # self.dateTimeEdit.setDisplayFormat('yyyy-MM-dd HH:mm:ss')
        # self.dateTimeEdit.setDateTime(QDateTime.currentDateTime())
        # self.dateTimeEdit.setCalendarPopup(True)

        # 设置时间格式
        self.timeEdit.setDisplayFormat('HH:mm:ss')
        self.lineEditDay.setText('0')

    '''
       函数名：ui_opera_select_all_device
       功能：界面默认值初始化
       参数：无
       返回值：无
       修改时间：
           2020年5月6日14:13:41
    '''
    def ui_opera_select_all_device(self):
        n = 0
        for table_row in range(10):  # 5->10
            for table_column in range(12):
                ip = self.tableWidget.item(table_row, table_column).text()
                if ip == "":
                    continue
                ip_current = {ip: "pass"}
                ip_online_dictionary.update(ip_current)
                self.tableWidget.item(table_row, table_column).setBackground(online_color)
                n += 1
        self.online_device_Edit.setText(str(n))

    '''
     函数名：ui_opera_config_refresh
     功能：
        读取配置文件
        初始化表格
     参数：无
     返回值：无
     修改时间：
         2020年5月11日09:42:03     
    '''

    def ui_opera_config_refresh(self):
        self.ui_read_config_init()
        self.ui_table_init()

    '''
    函数名：start_test
    功能：开始按钮动作
    参数：无
    返回值：无
    修改时间：
     2020年5月6日14:13:41
     2020年5月9日14:49:38
        #创建线程实现以下接口
        #添加标志位判断按钮是否点击
        #单击后设置不可点击
        #通过g_online_device判断ip是否存在，所以应该
    2020年5月12日12:02:27
        刷新界面配置
    '''

    def ui_opera_start(self):
        printf('ui_opera_start...')
        t, timeNum = self.getCounterTimeDaySec()
        if timeNum > 0:
            self.thread_1 = Thread_1()  # 创建线程
            self.thread_1.counter = timeNum
            self.thread_1.start()  # 开始线程
            self.thread_1.trigger.connect(self.fun_timer)
            return

        self.current_test_mod = self.comboBox.currentText()
        if self.current_test_mod == '设备ip筛选测试':
            print("你点了开始，选择了ip筛选检测")
            OnlineIpArry, num = self.ip_filter_v3()
            printf("OnlineIpArry:" + str(OnlineIpArry))
            # ret = self.MessageBoxDef("ip在线测试", str.format("ip在线测试完成，在线ip数为：%d" % num))
            # if ret == False:
            #     return
        # '热点-大电围',
        # '微热点-百才帮',
        # '微热点-超讯',
        elif self.current_test_mod == '热点-大电围':
            nu = self.test_num_Edit.text()
            if nu == "0":
                self.test_num_Edit.setText("1")
                nu = "1"

            number = int(nu)
            online_device = self.get_table_online_ip()
            online_deviceCheck = self.ip_filter_v4(online_device)
            # ret = self.MessageBoxDef("ip在线测试",
            #                          str.format("选择的ip数为：%d,在线ip数为：%d\n在线ip为：%s\n确定开始测试吗？" % (len(online_device),
            #                                                                                   len(online_deviceCheck),
            #                                                                                   online_deviceCheck)))
            # if ret == False:
            #     return

            # self.thread_creater_pa_mcb_kb_reboot(online_device, number)
            self.thread_creater_pa_mcb_kb_reboot_testAll(online_device, number,
                                                         self.thread_runner_pa_mcb_kb_reboot,
                                                         resultReportHeadLine)
            QMessageBox.information(self, "测试完成", str.format("测试完成"))

        elif self.current_test_mod == '微热点-百才帮':
            nu = self.test_num_Edit.text()
            if nu == "0":
                self.test_num_Edit.setText("1")
                nu = "1"

            number = int(nu)
            online_device = self.get_table_online_ip()
            online_deviceCheck = self.ip_filter_v4(online_device)
            # ret = self.MessageBoxDef("微热点-百才帮-主控,基带板断电重启",
            #                          str.format("选择的ip数为：%d,在线ip数为：%d\n在线ip为：%s\n确定开始测试吗？" % (len(online_device),
            #                                                                                   len(online_deviceCheck),
            #                                                                                   online_deviceCheck)))
            # if ret == False:
            #     return

            # self.thread_creater_pa_mcb_kb_reboot_baicaibang(online_device, number)
            self.thread_creater_pa_mcb_kb_reboot_testAll(online_device, number,
                                                         self.thread_runner_mcbcutdown_kb_reboot_baicaibang,
                                                         resultReportHeadLine_BCB)
            QMessageBox.information(self, "微热点-百才帮测试", str.format("测试完成"))
        elif self.current_test_mod == '微热点-超讯':
            nu = self.test_num_Edit.text()
            if nu == "0":
                self.test_num_Edit.setText("1")
            number = int(nu)
            online_device = self.get_table_online_ip()
            online_deviceCheck = self.ip_filter_v4(online_device)
            # ret = self.MessageBoxDef("微热点-超讯-主控,基带板软重启",
            #                          str.format("选择的ip数为：%d,在线ip数为：%d\n在线ip为：%s\n确定开始测试吗？" % (len(online_device),
            #                                                                                   len(online_deviceCheck),
            #                                                                                   online_deviceCheck)))
            # if ret == False:
            #     return

            # self.thread_creater_pa_mcb_kb_reboot_chaoxun(online_device, True)
            self.thread_creater_pa_mcb_kb_reboot_testAll(online_device, number,
                                                         self.thread_runner_mcbsoft_kb_reboot_chaoxun,
                                                         resultReportHeadLine_CX)

            QMessageBox.information(self, "微热点-超讯-主控重启，基带板软重启测试", str.format("测试完成"))
        elif self.current_test_mod == '主控重启':
            self.ui_opera_mcb_reboot()

        elif self.current_test_mod == 'GPIO断电重启':
            self.ui_opera_gpio_power_down_reboot()

    '''
         函数名：stop_test
         功能：停止按钮动作
         参数：无
         返回值：无
         修改时间：
             2020年5月6日14:13:41
    '''

    def ui_opera_stop(self):
        # 取得当前的计时时间：s
        # t, timeNum = self.getCounterTimeDaySec()
        # # 创建线程
        # self.thread_1 = Thread_1()
        # # 告诉线程计时多少秒
        # self.thread_1.counter = timeNum
        # self.thread_1.start()  # 开始线程
        # self.thread_1.trigger.connect(self.fun_timer)

        # login("http://192.168.3.201")
        # _thread.start_new_thread(self.print_time, ("Thread-1", 2,))
        # _thread.start_new_thread(self.ui_opera_start)
        online_device = self.get_table_online_ip()
        self.thread_creater_pa_mcb_kb_reboot_testAll(online_device,2, self.thread_runner_pa_mcb_kb_reboot,resultReportHeadLine)
        pass

    def print_time(self,threadName, delay):
        count = 0
        while count < 50:
            time.sleep(delay)
            count += 1
            print("%s: %s" % (threadName, time.ctime(time.time())))

    def fun_timer(self):
        print(time.strftime('%Y-%m-%d %H:%M:%S'))
        t = self.timeEdit.time()
        t1,timeNum =self.getCounterTimeSec()
        day = int(self.lineEditDay.text())
        printf('day:' + str(day))
        printf('timeNum:'+str(timeNum))
        if day > 0 and timeNum == 0:
            self.timeEdit.setTime(t.addSecs(-1))
            self.lineEditDay.setText(str(day-1))
            return
        elif day == 0 and timeNum > 1:
            self.timeEdit.setTime(t.addSecs(-1))
            return
        elif day > 0 and timeNum > 1:
            self.timeEdit.setTime(t.addSecs(-1))
            return
        elif day == 0 and timeNum == 1:
            self.timeEdit.setTime(t.addSecs(-1))
            self.ui_opera_start()
        else:
            self.timeEdit.setTime(t.addSecs(-1))
            self.ui_opera_start()

        # if timeNum >= 1:
        #     self.timeEdit.setTime(t.addSecs(-1))
        # else:
        #     self.timeEdit.setTime(t.addSecs(-1))
        #     self.ui_opera_start()

    def getCounterTimeDaySec(self):
        t = self.timeEdit.time()
        day = int(self.lineEditDay.text())
        return t, t.second() + t.minute() * 60 + t.hour() * 60 * 60 + day*24*60*60

    def getCounterTimeSec(self):
        t = self.timeEdit.time()
        return t, t.second() + t.minute() * 60 + t.hour() * 60 * 60

    '''
    函数名：ui_opera_pa_switch
    功能：功放开关操作
    参数：无
    返回值：无
    修改时间：
        2020年5月6日14:13:41
    '''

    def ui_opera_pa_switch(self):
        self.pa_opera = False
        if self.pa_checkBox.isChecked():  # open
            self.pa_opera = True
        else:  # close
            pass

        print("你点了开始，选择了功放开关测试")
        nu = self.test_num_Edit.text()
        if nu == "0":
            self.test_num_Edit.setText("1")

        online_device = self.get_table_online_ip()
        online_deviceCheck = self.ip_filter_v4(online_device)
        ret = self.MessageBoxDef("功放开关测试",
                                 str.format("选择的ip数为：%d,在线ip数为：%d\n在线ip为：%s\n确定开始测试吗？" % (len(online_device),
                                                                                          len(online_deviceCheck),
                                                                                          online_deviceCheck)))
        if ret == False:
            return

        self.thread_creater_pa_switch(online_deviceCheck, self.pa_opera)
        pa_false = []
        for ip in online_device:
            if pa_opera_dict.get(ip) == "false":
                pa_false.append(ip)

        if len(pa_false) == 0:
            QMessageBox.information(self, "功放开关测试", str.format("测试完成,"))
        else:
            QMessageBox.information(self, "功放开关测试", str.format("测试完成,失败：%s" % pa_false))

    '''
    函数名：mcb_reboot_operate
    功能：循环的执行主控重启，每五分钟一次
    参数：times：重启次数
    返回值：无
    修改时间：
        2020年5月29日14:53:17
    '''

    def ui_opera_mcb_reboot(self):
        nu = self.test_num_Edit.text()
        if nu == "0":
            self.test_num_Edit.setText("1")

        online_device = self.get_table_online_ip()
        online_deviceCheck = self.ip_filter_v4(online_device)
        ret = self.MessageBoxDef("主控重启测试",
                                 str.format("选择的ip数为：%d,在线ip数为：%d\n在线ip为：%s\n确定开始测试吗？" % (len(online_device),
                                                                                          len(online_deviceCheck),
                                                                                          online_deviceCheck)))
        if ret == False:
            return

        self.thread_creater_mcb_reboot(online_deviceCheck, True)

        QMessageBox.information(self, "主控重启测试", str.format("主控重启测试完成"))

    '''
      函数名：ui_opera_gpio_power_down_reboot
      功能：
        通过3160控制板的断电重启操作，进行多快板子的断电重启测试
        3160 ip:192.168.3.11
        单线程
        输出结果报告
        
        
      参数：无
      返回值：无
      修改时间：
        2020年6月4日17:18:41
    '''

    def ui_opera_gpio_power_down_reboot(self):
        nu = self.test_num_Edit.text()
        if nu == "0":
            self.test_num_Edit.setText("1")
        # ["ip","id","重启时间","上线时间","结果"],
        gpio_pdr_result = (
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
            ["", "", "", "", ""],
        )

        online_device = self.get_table_online_ip()
        online_deviceCheck = self.ip_filter_v4(online_device)
        ret = self.MessageBoxDef("gpio断电重启测试",
                                 str.format("选择的ip数为：%d,在线ip数为：%d\n在线ip为：%s\n确定开始测试吗？" % (len(online_device),
                                                                                          len(online_deviceCheck),
                                                                                          online_deviceCheck)))
        if ret == False:
            return

        ############################################################
        self.MCB_REBOOT_TABLE_NAME = ["gpio_power_down_reboot", ]
        self.MCB_REBOOT_REPORT_FILE_NAME = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))

        self.REPORT_PATH = self.reportLineEdit.text()
        if os.path.exists(self.REPORT_PATH) == False:
            os.mkdir(self.REPORT_PATH)  # 目录不存在，创建
        else:
            pass

        try:
            self.openpyxl_create_report_new(
                self.REPORT_PATH,
                self.MCB_REBOOT_REPORT_FILE_NAME,
                self.MCB_REBOOT_TABLE_NAME)

            self.openpyxl_addin_new(
                self.REPORT_PATH,
                self.MCB_REBOOT_REPORT_FILE_NAME,
                self.MCB_REBOOT_TABLE_NAME[0],
                GPIO_POWER_DOWN_REBOOT_HEADLINE,
                0)
        except Exception as err:
            print(err)
            return

        # 循环测试
        '''
        #联通测试
        for ip in online_deviceCheck:
            if ip is None:
                continue

            url = "http://" + ip + "/#/login/"
            dev_id,err = login_get_device_id(url)
            if err == 0:
                pass
        '''
        printf("开始循环测试")
        for i in range(int(nu)):
            rebootTime = gpio_power_down_reboot("http://192.168.3.11/#/login/", band_1)
            if rebootTime == "":
                printf("gpio断电重启失败，等待一会在尝试")
                time.sleep(10)
                continue

            printf("等待3分钟后进行核心板登录测试")
            time.sleep(60 * 3)
            arrIndex = 0

            for ip in online_deviceCheck:
                if ip is None:
                    continue
                gpio_pdr_result[arrIndex][0] = ip
                gpio_pdr_result[arrIndex][1] = None
                gpio_pdr_result[arrIndex][2] = rebootTime
                gpio_pdr_result[arrIndex][3] = None
                gpio_pdr_result[arrIndex][4] = "false"
                url = "http://" + ip + "/#/login/"
                dev_id, err = login_get_device_id(url)
                if err == -1:
                    printf("基带板重启失败")
                    gpio_pdr_result[arrIndex][4] = "false"
                else:
                    printf("基带板重启成功")
                    gpio_pdr_result[arrIndex][1] = dev_id
                    gpio_pdr_result[arrIndex][3] = get_current_time_str()
                    gpio_pdr_result[arrIndex][4] = "pass"

                arrIndex += 1

            # 记录报告，结果
            for val in gpio_pdr_result:
                print("val[0]:%s" % val[0])
                if val[0] is None or val[0] == "":
                    continue
                self.openpyxl_addin_new(
                    self.REPORT_PATH,
                    self.MCB_REBOOT_REPORT_FILE_NAME,
                    self.MCB_REBOOT_TABLE_NAME[0],
                    val,
                    1)
            printf("等待2分钟")
            time.sleep(2 * 60)

        ############################################################

        QMessageBox.information(self, "gpio断电重启测试", str.format("gpio断电重启测试完成"))

    '''
           函数名：chose_table_cell
           功能：
               获取当前在线设备数值
               获取单元格行列值
               获取单元格文本:ip
               判断该ip在字典中的属性值，"pass" or "false"
               更新字典属性值
               更新单元格颜色
               更新在线设备数
           参数：无
           返回值：无
           修改时间：
           2020年5月6日14:13:41
           2020年5月11日09:35:50
       '''

    def ui_opera_chose_table_cell(self, index):
        online_dev_num = int(self.online_device_Edit.text())
        table_column = index.column()
        table_row = index.row()
        ip = self.tableWidget.item(table_row, table_column).text()
        val = ip_online_dictionary.get(ip)  # 查询字典中ip的属性是 pass 还是 false
        if val == "pass":
            ip_current = {ip: "false"}
            ip_online_dictionary.update(ip_current)
            self.tableWidget.item(table_row, table_column).setBackground(offline_color)
            if online_dev_num > 0:
                self.online_device_Edit.setText(str(online_dev_num - 1))
                print("你取消了(%d,%d)" % (table_column + 1, table_row + 1))

        elif val == "false":
            ip_current = {ip: "pass"}
            ip_online_dictionary.update(ip_current)
            self.tableWidget.item(table_row, table_column).setBackground(online_color)
            self.online_device_Edit.setText(str(online_dev_num + 1))
            print("你选择了(%d,%d)" % (table_column + 1, table_row + 1))

    '''
        函数名：chose_driver_path
        功能：选择驱动路径
        参数：无
        返回值：无
        修改时间：
            2020年5月6日14:13:41
    '''

    def ui_opera_chose_driver_path(self):
        # self.config_create_data()
        log_path = QFileDialog.getExistingDirectory() + "chromedriver.exe"
        self.chromeDriverDirLineEdit.setText(log_path)
        pass

    '''
         函数名：chose_log_path
         功能：选择日志保存路径
         参数：无
         返回值：无
         修改时间：
             2020年5月6日14:13:41
    '''

    def ui_opera_chose_log_path(self):
        log_path = QFileDialog.getExistingDirectory()
        self.logDirLineEdit.setText(log_path)
        pass

    '''
         函数名：chose_report_path
         功能：选择报告保存路径
         参数：无
         返回值：无
         修改时间：
             2020年5月6日14:13:41
    '''

    def ui_opera_chose_report_path(self):
        report_path = QFileDialog.getExistingDirectory()
        self.reportLineEdit.setText(report_path)
        pass


if __name__ == '__main__':
    app = QApplication(sys.argv)
    myWin = MyWindow()
    myWin.show()
    sys.exit(app.exec_())
