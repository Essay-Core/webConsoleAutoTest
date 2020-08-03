# 标准库
import time
from selenium import webdriver

# 第三方库
from PyQt5.QtGui import QBrush, QColor
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font


# 单元格颜色填充选项
red_fill = PatternFill("solid", fgColor="FF0000")
green_fill = PatternFill("solid", fgColor="00FF00")

# 字体格式，颜色和大小
font_pass = Font(size=16, bold=True, color="00FF00")
font_false = Font(size=16, bold=True, color="FF0000")

online_color = QBrush(QColor(95, 215, 169))
offline_color = QBrush(QColor(255, 0, 0))

Chrome_path = r"./chromedriver.exe"

# 先定位主筐体的位置，让后通过相对位置来定位表格中的元素
main_xpath = '//main[@class="el-main main"]'

# 登录界面
inputuser_box = '//input[@name="username"]'
inputpasswd_box = '//input[@name="password"]'
login_submit_btn = '//span[text()="登录"]'

# 设备管理
screenfull = '//div[@class="el-tooltip hamburger-container"]'
dev_manage = "//span[text()='设备管理']"
zm_dev_manage = "//span[text()='侦码设备管理']"
reflash = '//i[@class="el-icon-refresh"]'

dev_msg_current_time = '#pane-device-info > div > div:nth-child(2) > fieldset:nth-child(1) > div:nth-child(2) > div.el-col.el-col-9'
dev_msg_version = '#pane-device-info > div > div:nth-child(2) > fieldset:nth-child(2) > div:nth-child(2)'
dev_msg_run_time = '#pane-device-info > div > div:nth-child(2) > fieldset:nth-child(3) > div'
ssd_mess = '#pane-device-info > div > div:nth-child(2) > fieldset:nth-child(2) > div:nth-child(2)'  # 磁盘信息
close_bt = '#app > div > div.main-container > section > div > div > div > div.el-dialog__header > button'  # 关闭按钮


dev_line_bt = '(//button[@class="el-button el-button--primary el-button--mini"])[1]'
dev_id = '//span[text()="190829"]'
dev_online = '//span[text()="在线"]'
dev_config_bt = '//span[text()="在线"]'

# 设备ID页面
dev_reflash = '(//span[text()="刷新"])[1]'

def openpyxl_create_report_new(filePath, fileName, tableNames):
    path = filePath + fileName + ".xlsx"
    data = openpyxl.Workbook()  # 新建工作簿
    i = 0
    for name in tableNames:
        print("create table name:%s" % tableNames)
        data.create_sheet(name, i)  # 添加页,设置第一个工作页
        i += 1
    data.save(path)  # 一定要保存


def openpyxl_addin_new(filePath, fileName, tableName, values, IfHead):
    fileName = filePath + fileName + ".xlsx"
    data = openpyxl.load_workbook(fileName)
    table = data[tableName]
    nrows = table.max_row  # 获得行数+1的值，即下一行应该写入的地方
    if IfHead == 0:
        pass  # 此时为标题头
    else:
        nrows += 1
    col = 1
    resultFlags = 0
    for value in values:
        print("value:%s,row:%d,col:%d" % (value, nrows, col))
        if value is None:
            col += 1
            continue
        table.cell(row=nrows, column=col).value = value
        if 'pass' == value:
            # 设置字体颜色
            table.cell(row=nrows, column=col).fill = green_fill
        elif 'false' == value:
            resultFlags += 1
            table.cell(row=nrows, column=col).fill = red_fill
        # 设置居中对齐
        table.cell(row=nrows, column=col).alignment = Alignment(horizontal='center', vertical='center')
        col += 1
    print(fileName)
    data.save(fileName)


def get_current_time_str():
    update_str = time.strftime('%Y-%m-%d %H:%M:%S')
    return update_str


USER_NAME = 'admin'
PASS_WORD = '123456'

def login(url):
    driver = webdriver.Chrome(Chrome_path)
    driver.implicitly_wait(10)
    try:
        driver.get(url)
        driver.maximize_window()
        driver.find_element_by_xpath(inputuser_box).send_keys(USER_NAME)
        driver.find_element_by_xpath(inputpasswd_box).send_keys(PASS_WORD)
        driver.find_element_by_xpath(login_submit_btn).click()
        time.sleep(20)
        return driver, 1
    except Exception as err:
        print(err)
        driver.quit()
        return driver, 0

REPORT_FILE_PATH =  '../report/'
REPORT_FILE_NAME = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
REPORT_TABLE_NAME = ['zm_dev_manage',]
# 测试结果
ZM_DEV_MANAGE_HEADLINE = [
    "设备ID",
    "设备当前设备时间",
    "设备版本信息",
    "设备运行时长",

    "核心板名称",
    "核心板工作状态",

    "功放名称",
    "功放开关状态",

    "测试结果"
]
def zm_device_managent(url):
    zdmStatus = ['', '','', '','', '','', '','']

    openpyxl_create_report_new(REPORT_FILE_PATH, REPORT_FILE_NAME, REPORT_TABLE_NAME)
    openpyxl_addin_new(REPORT_FILE_PATH,REPORT_FILE_NAME,REPORT_TABLE_NAME[0],ZM_DEV_MANAGE_HEADLINE,0)
    driver,loginStatus = login(url)
    if loginStatus == 1:
        print('login success')
    else:
        time.sleep(10)
        driver, loginStatus = login(url)


    driver.find_element_by_xpath(screenfull).click()
    time.sleep(1)
    print('click device manage')
    driver.find_element_by_xpath(dev_manage).click()
    time.sleep(1)
    print('click zm device mange')
    driver.find_element_by_xpath(zm_dev_manage).click()
    time.sleep(10)
    for i in range(3000):
        for i in range(50):
            try:
                main_xpath = '//div[@class="pagination-container"]'
                locater = driver.find_element_by_xpath(main_xpath)
                locater.find_element_by_xpath(reflash).click()  # 刷新
                print("click reflash")
                time.sleep(10)

                # 设备信息页面
                main_xpath = '//div[@class="el-table el-table--fit el-table--striped el-table--border el-table--scrollable-y el-table--enable-row-transition el-table--medium"]'
                dev_id = '//*[@id="app"]/div/div[2]/section/div/div/div/div[1]/span'
                locater = driver.find_element_by_xpath(main_xpath)
                locater.find_element_by_xpath(dev_line_bt).click()  # 定位可点击按钮进入设备信息界面
                print('click device configure button')
                time.sleep(10)
                id = driver.find_element_by_xpath(dev_id).text
                print('get device ID ')
                s = id[len('设备ID：'):]
                if s != '190829':
                    pass
                zdmStatus[0] = s
                print("device ID is %s"%id)
                break
            except Exception as err:
                print(err)

        try:
            current_time = driver.find_element_by_css_selector(dev_msg_current_time).text
            print(current_time)
            zdmStatus[1] = current_time
            version = driver.find_element_by_css_selector(dev_msg_version).text
            print(version)
            zdmStatus[2] = version
            run_time = driver.find_element_by_css_selector(dev_msg_run_time).text
            print(run_time)
            zdmStatus[3] = run_time
        except Exception as err:
            print(err)

        kb_msg_page = '//div[@aria-controls="pane-core-board-info"]'
        bk_msg_name = '#pane-core-board-info > div > div:nth-child(2) > div > div.el-col.el-col-10 > fieldset:nth-child(1) > span'
        pa_name = '#pane-core-board-info > div > div:nth-child(2) > div > div.el-col.el-col-14 > fieldset:nth-child(1) > span > b'
        pa_open_status = '//span[@class="el-switch__label el-switch__label--right is-active"]'  # 功放开启时可定位成功

        try:
            driver.find_element_by_xpath(kb_msg_page).click()
            time.sleep(5)

            text = driver.find_element_by_css_selector(bk_msg_name).text
            print(text)
            zdmStatus[4] = text
            text = driver.find_element_by_css_selector(pa_name).text
            print(text)
            zdmStatus[6] = text
        except Exception as err:
            print(err)

        '''
         try:
            driver.find_element_by_css_selector(pa_open_status)
            print("功放开启")
        except Exception as err:
            print(err)
            print("功放无法定位或已关闭状态")
        '''
        # 关闭结束此轮测试
        driver.find_element_by_css_selector(close_bt).click()
        openpyxl_addin_new(REPORT_FILE_PATH, REPORT_FILE_NAME, REPORT_TABLE_NAME[0], zdmStatus, 1)


if __name__ == '__main__':
    url = 'http://192.168.3.234/lbsweb'
    zm_device_managent(url)

